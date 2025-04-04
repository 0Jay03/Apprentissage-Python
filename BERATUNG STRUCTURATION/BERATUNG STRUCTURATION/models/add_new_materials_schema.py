import json
import re

import openpyxl

# Load the JSON data with UTF-8 encoding
with open('data_to_add.json', 'r', encoding='utf-8') as file:
    products = json.load(file)

# Open the main Excel file with openpyxl
wb = openpyxl.load_workbook('PRODUKTLISTE.xlsx')
ws = wb.active

# Open the additional Excel file for cost groups
wb_costs = openpyxl.load_workbook('Eingabehilfe.xlsx')
ws_costs = wb_costs.active

# Prepare a list to store the results
results = []

def extract_numeric_prefix_regex(text):
    if not text.strip():  # Check if the text is empty or whitespace
        return None
    # Use a regex pattern to find the first occurrence of one or more numbers separated by dots
    match = re.search(r'\d+(\.\d+)*', text)
    if match:
        return match.group(0).strip()  # Return the matched group, stripped of any surrounding whitespace
    return None  # Return None if no match is found



# Process each product in the JSON file
for product in products:
    name = str(product['name']).strip().lower()  # Normalize to lowercase and strip whitespace
    manufacturer = str(product.get('manufacturerName', '')).strip().lower()

    # Search for the corresponding row in the main Excel
    found_row = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        cell_value = str(row[2]).strip().lower() if row[2] is not None else ''
        if cell_value == name:
            found_row = row
            break

    # Search for cost group information in the additional Excel
    cost_group2 = product['costGroups2']
    cost_group3 = product['costGroups3']

    # Extract the necessary information if found
    if found_row:
        application_area = found_row[3] if found_row[3] is not None else ""
        application_field_position = found_row[4] if found_row[4] is not None else ""
        verification_document = found_row[6] if found_row[6] is not None else ""
        company = found_row[10] if found_row[10] is not None else ""

        # Extract concerned employees
        employees_concerned = []
        for i, cell in enumerate(found_row[15:31], start=15):
            if cell == 'ja':
                employees_concerned.append(ws[1][i].value.strip())

                # Structure the data to be saved in the new JSON file
        output_data = {
            'applicationArea': application_area,
            'verificationDocument': verification_document,
            'manufacturerName': str(product.get('manufacturerName', '')).strip(),
            'name': str(product['name']).strip(),
            'costGroups2': cost_group2,
            'costGroups3': cost_group3,
            'productLinks': [],
            'quantity': "Nicht spezifiziert",
            'useArea': "Nicht spezifiziert",
            'company': company,
            'employeesConcerned': employees_concerned,
            'compartmentsConcerned': [],
            'applicationFieldPosition': extract_numeric_prefix_regex(application_field_position)
        }

        results.append(output_data)
    else:
        print(f"⚠️ Le nom '{name}' n'est pas trouvé dans le fichier Excel principal.")

# Save the results in a new JSON file with UTF-8 encoding
with open('data_to_add2.json', 'w', encoding='utf-8') as outfile:
    json.dump(results, outfile, ensure_ascii=False, indent=4)

print(f"✅ Successfully processed {len(results)} items and saved them in data_to_add2.json")